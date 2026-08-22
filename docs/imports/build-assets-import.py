# -*- coding: utf-8 -*-
# مولّد ملف استيراد الأصول من ملفات الحصر الميداني
#
# المشكلة: ملف الحصر مقسّم إلى ١٢ ورقة حسب الجهة، وأعمدته لا تطابق أعمدة قالب
# الاستيراد في النظام؛ كما أنّ عمود «عمود1» فيه يحمل معنيين: ماركةَ الجهاز في
# صفوف التكييف، وعددَ الوحدات في غيرها.
#
# المبدأ: التحويل قاعدةٌ مكتوبة لا اجتهادٌ يدوي — جدولُ أصناف يربط كلَّ مسمّى
# ميداني بنوعٍ ونوعٍ فرعيّ من قوائم النظام (ASSET_TYPES_LIST / ASSET_SUBTYPES في
# index.html)، وجدولُ مبانٍ يربط اسم الجهة بالمبنى المعرَّف في BUILDINGS.
# فمسمّى غيرُ معروف يُبلَّغ عنه ولا يُخمَّن.
#
# القرارات:
#   • البند ذو الكمية n يُفرد في n صفوف — كلُّ جهازٍ فعليّ أصلٌ مستقلّ في السجل،
#     وإلا تعذّر ربطُ أمر عملٍ بجهازٍ بعينه. ويُرقَّم في الملاحظات (وحدة i من n).
#   • «وكالة التحويل الرقمي» ← «مبنى تقنية المعلومات» (أقربُ مبنًى معرَّف).
#   • «الوثائق والمحفوظات» و«الوحدة المركزية لاعتماد المخططات» يُكتبان باسمَيهما
#     ويحتاجان إضافةً من الإعدادات ← المباني قبل الاستيراد.
#   • صفوف الباركينج والبوابات كلُّها تحت «مبنى الأمانة الرئيسي»، واسمُ البوابة
#     في الموقع الفرعي.
#
# التشغيل: عدِّل SRC ليشير إلى ملف الحصر ثم:  python3 docs/imports/build-assets-import.py
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = '/root/.claude/uploads/6ce28432-95eb-5bd2-8954-4579d0bbfc10/6c62d457-_________________.xlsx'
OUT = '/tmp/claude-0/-home-user-Fast-buildings-maintenance-/6ce28432-95eb-5bd2-8954-4579d0bbfc10/scratchpad/قالب_استيراد_الأصول_معبأ.xlsx'

BUILDING_MAP = {
 'وكالة التحويل الرقمي':'مبنى تقنية المعلومات',
 'وكالة الخدمات':'مبنى وكالة الخدمات',
 'الجودة والأمتثال':'مبنى الجودة والامتثال',
 'الأراضي والتخطيط':'مبنى الأراضي',
 'الوثائق والمحفوظات':'مبنى الوثائق والمحفوظات',
 'الوحدة المركزية لأعتماد المخططات':'مبنى الوحدة المركزية لاعتماد المخططات',
 'خدمات المستفيدين':'مبنى خدمة المستفيدين',
 'الأمن السبرانى':'مبنى الأمن السيبراني',
 'الأمن والسلامة':'مبنى الأمن والسلامة',
 'المراجعة الداخلية':'مبنى المراجعة الداخلية',
 'مبنى الأمانة الرائيسي':'مبنى الأمانة الرئيسي',
}
# (نوع الجهاز, النوع الفرعي, الموديل, وصف إضافي للملاحظات)
ITEM_MAP = {
 'مكيف اسبيلت':('مكيف','Split','',''), 'مكيف اسبليت':('مكيف','Split','',''),
 'مكبف اسبيلت':('مكيف','Split','',''), 'مكيف كاست':('مكيف','Cassette','',''),
 'خزان مياه علوي':('خزان مياه','Overhead','',''), 'خزان مياة علوي':('خزان مياه','Overhead','',''),
 'خزان مياه أرضي':('خزان مياه','Underground','',''), 'خزان مياه ارضي':('خزان مياه','Underground','',''),
 'خزان مياه حريق ارضي':('خزان مياه','Underground','','خزان مياه حريق'),
 'خزان ديزل':('معدة أخرى','أخرى','','خزان ديزل لتغذية المولد'),
 'ماتور 1حصان':('مضخة','Centrifugal','1 حصان',''), 'ماتور 1 حصان':('مضخة','Centrifugal','1 حصان',''),
 'ماتور2\\1 حصان':('مضخة','Centrifugal','1/2 حصان',''), 'ماتور 2\\1 حصان':('مضخة','Centrifugal','1/2 حصان',''),
 'غطاس 1 حصان ارضي':('مضخة','Submersible','1 حصان','غطاس أرضي'),
 'مضخة مياه كبيرة':('مضخة','Centrifugal','','مضخة مياه كبيرة'),
 'مجموعة مضخات حريق':('مضخة','Fire Pump','','مجموعة مضخات حريق'),
 'لوحة كهرباء رائيسية':('لوحة كهربائية','Main Panel','',''),
 'لوحة رائيسية كبيرة':('لوحة كهربائية','MDB','','لوحة رئيسية كبيرة'),
 'لوحة كهرباء توزيع':('لوحة كهربائية','DB','',''),
 'لوحة كهرباء انارة':('لوحة كهربائية','Sub Panel','','لوحة إنارة'),
 'لوحة كهرباء افياش':('لوحة كهربائية','Sub Panel','','لوحة أفياش'),
 'لوحة كهرباء طرمبات البمب':('لوحة كهربائية','Sub Panel','','لوحة طرمبات غرفة المضخات'),
 'لوحة سمبيت':('لوحة كهربائية','Sub Panel','','لوحة سمبيت'),
 'لوحة البمب روم':('لوحة كهربائية','Sub Panel','','لوحة غرفة المضخات (Pump Room)'),
 'لوحة كهرباء التكييف':('لوحة كهربائية','Sub Panel','','لوحة تغذية التكييف'),
 'لوحة مروحة شفط':('لوحة كهربائية','Sub Panel','','لوحة مراوح الشفط'),
 'لوحة كهرباء للمصاعد':('لوحة كهربائية','Sub Panel','','لوحة تغذية المصاعد'),
 'مولد كهرباء':('مولد كهربائي','Diesel','',''),
 'بوابة كهربائية':('معدة أخرى','أخرى','','بوابة كهربائية'),
}
FLOOR_MAP = {'السطح':'السطح','البدروم':'البدروم','الأرضي':'الدور الأرضي','الارضى':'الدور الأرضي',
 'الأول':'الدور الأول','الثاني':'الدور الثاني','الثالث':'الدور الثالث','الرابع':'الدور الرابع',
 'الخامس':'الدور الخامس','السادس':'الدور السادس','السابع':'الدور السابع','الثامن':'الدور الثامن'}
GATE_MAP = {'البوابة الرائيسية':'البوابة الرئيسية','بوابة 2':'بوابة 2','بوابة 3':'بوابة 3','الباركينج':'الباركينج'}

def s(v): return '' if v is None else str(v).strip()

wb = openpyxl.load_workbook(SRC, data_only=True)
rows_out, problems = [], []

for ws in wb.worksheets:
    parking = (ws.title == 'حصر الباركينج')
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7, values_only=True):
        item = s(r[1])
        if not item: continue
        if parking:
            qty_raw, bld_raw, sub_raw = s(r[2]), s(r[3]), s(r[4])
            floor_raw = ''
        else:
            bld_raw, sub_raw, qty_raw, floor_raw = s(r[2]), s(r[3]), s(r[4]), s(r[5])
        if item not in ITEM_MAP:
            problems.append(f'{ws.title}: صنف غير معروف "{item}"'); continue
        typ, sub, model, extra = ITEM_MAP[item]

        # المبنى + الموقع الفرعي
        if parking:
            building = 'مبنى الأمانة الرئيسي'
            sub_location = GATE_MAP.get(bld_raw, bld_raw) if bld_raw not in BUILDING_MAP else 'الباركينج'
            floor = 'الباركينج — الساحات الخارجية'
        else:
            if bld_raw not in BUILDING_MAP:
                problems.append(f'{ws.title}: مبنى غير معروف "{bld_raw}"'); continue
            building = BUILDING_MAP[bld_raw]
            sub_location = 'مرافق المبنى العامة' if sub_raw == bld_raw else sub_raw
            floor = FLOOR_MAP.get(floor_raw, floor_raw)
            if floor_raw and floor_raw not in FLOOR_MAP:
                problems.append(f'{ws.title}: دور غير معروف "{floor_raw}"')

        # عمود1 = ماركة للمكيفات، وكمية لغيرها
        qty = 1
        if typ == 'مكيف':
            if qty_raw: model = qty_raw
        elif qty_raw:
            m = re.match(r'^\d+', qty_raw)
            if m: qty = int(m.group(0))
            else: problems.append(f'{ws.title}: كمية غير رقمية "{qty_raw}" للصنف {item}')

        def _n(x):
            x = re.sub(r'[ًٌٍَُِّْ\s]', '', x)
            for a,b in [('أ','ا'),('إ','ا'),('آ','ا'),('ة','ه'),('ى','ي')]: x = x.replace(a,b)
            return x.replace('كهرباء','')
        keep_extra = bool(extra) and _n(extra) not in _n(item)
        base_note = ' — '.join([item] + ([extra] if keep_extra else []))
        for i in range(1, qty+1):
            note = f'{base_note} — مصدر: {ws.title}'
            if qty > 1: note += f' (وحدة {i} من {qty})'
            rows_out.append([typ, sub, sub_location, floor, building, '', model, '', '', note])

# ── كتابة الملف بنفس ترويسة القالب ──
HEAD = ["نوع الجهاز *","النوع الفرعي","الموقع الفرعي","الدور","المبنى *","الرقم التسلسلي","الموديل",
        "تاريخ التركيب (YYYY-MM-DD)","انتهاء الضمان (YYYY-MM-DD)","ملاحظات"]
out = openpyxl.Workbook(); o = out.active; o.title = 'الأصول'; o.sheet_view.rightToLeft = True
o.append(HEAD)
for row in rows_out: o.append(row)

hf = PatternFill('solid', fgColor='1F4E79'); thin = Side(style='thin', color='D0D7DE')
for c in o[1]:
    c.font = Font(bold=True, color='FFFFFF', size=11); c.fill = hf
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for i,w in enumerate([16,14,34,24,32,16,14,20,20,52],1):
    o.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
o.row_dimensions[1].height = 32
o.freeze_panes = 'A2'
o.auto_filter.ref = f'A1:J{o.max_row}'
for row in o.iter_rows(min_row=2, max_row=o.max_row, max_col=10):
    for c in row:
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
out.save(OUT)

print('صفوف مكتوبة:', len(rows_out))
from collections import Counter
for k,v in Counter(r[4] for r in rows_out).most_common(): print(f'  {k}: {v}')
print('\nنوع الجهاز:')
for k,v in Counter(r[0] for r in rows_out).most_common(): print(f'  {k}: {v}')
print('\nملاحظات/مشاكل:', len(problems))
for p in sorted(set(problems)): print('  -', p)
